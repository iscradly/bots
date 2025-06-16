import telebot
from telebot import types
import openpyxl
from datetime import datetime, timedelta

TOKEN = ""

bot = telebot.TeleBot(TOKEN)

# Устанавливаем команды для бота
bot.set_my_commands([
    telebot.types.BotCommand("start", "Начать работу")
])

# Функция для загрузки графиков из Excel
def load_schedule(filename='Grafiki.xlsx'):
    schedule_data = {}
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active

    # Предполагается, что первая колонка - ФИО, остальные - даты с графиками
    for row in sheet.iter_rows(min_row=2, values_only=True):
        fio = row[0]
        schedule_data[fio] = {sheet.cell(row=1, column=i + 1).value: row[i] for i in range(1, len(row))}

    return schedule_data

# Загружаем графики из файла при старте бота
schedule_data = load_schedule()

@bot.message_handler(commands=['start'])
def start(message):
    """Обрабатывает команду /start"""
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    btn1 = types.KeyboardButton('Посмотреть график работы на месяц')
    btn2 = types.KeyboardButton('Выбрать дату')
    markup.add(btn1, btn2)
    bot.send_message(message.chat.id, "Добрый день! Выберите действие", reply_markup=markup)

@bot.message_handler(func=lambda message: message.text == 'Посмотреть график работы на месяц')
def handle_month_schedule(message):
    bot.send_message(message.chat.id, 'Введите Фамилию и Имя полностью')

@bot.message_handler(func=lambda message: message.text == 'Выбрать дату')
def handle_date_selection(message):
    bot.send_message(message.chat.id, 'Введите дату в формате ДД.ММ')

@bot.message_handler(func=lambda message: True)
def handle_text(message):
    """Обрабатывает ввод ФИО сотрудника и возвращает график на месяц или на конкретную дату"""
    text = message.text.strip()

    if text in schedule_data:
        month_schedule = get_month_schedule(text)
        markup = types.ReplyKeyboardRemove()
        bot.send_message(message.chat.id, month_schedule, reply_markup=markup)
    elif '.' in text:  # Проверяем, введена ли дата
        date = text
        fio_schedule = get_schedule_by_date(date)
        if fio_schedule:
            bot.send_message(message.chat.id, fio_schedule)
        else:
            bot.send_message(message.chat.id, "На эту дату нет данных.")
    else:
        bot.send_message(message.chat.id, "Сотрудник не найден. Пожалуйста, проверьте ФИО.")

def get_month_schedule(fio):
    """Возвращает график на месяц для выбранного сотрудника"""
    today_date = datetime.now()
    month_schedule = []

    for i in range(31):  # Проходим по дням
        date = (today_date + timedelta(days=i)).strftime("%d.%m")  # Дата в формате ДД.ММ
        status = schedule_data[fio].get(date, 'Нет данных')
        month_schedule.append(f"{date}: {status}")

    return f"График для {fio} на месяц:\n" + "\n".join(month_schedule)

def get_schedule_by_date(date):
    """Возвращает график на конкретную дату для всех сотрудников"""
    schedule = []
    for fio, dates in schedule_data.items():
        status = dates.get(date, 'Нет данных')
        schedule.append(f"{fio}: {status}")
    return "График на " + date + ":\n" + "\n".join(schedule)

if __name__ == "__main__":
    bot.polling(none_stop=True)
